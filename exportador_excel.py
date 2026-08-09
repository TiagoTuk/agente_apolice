"""
Gera a planilha final com os resultados de todos os clientes processados.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import ARQUIVO_EXCEL_SAIDA, CAMPOS_EXTRACAO
from logger import get_logger

log = get_logger("exportador_excel")

COLUNAS_FIXAS = ["Cliente", "Ano", "Arquivo", "Caminho completo", "Origem", "Status", "Erro Detalhe"]


def exportar(resultados: list, caminho_saida: Path = ARQUIVO_EXCEL_SAIDA) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apólices"

    colunas_campos = list(CAMPOS_EXTRACAO.keys())
    cabecalho = COLUNAS_FIXAS + [c.replace("_", " ").title() for c in colunas_campos]
    ws.append(cabecalho)

    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    fundo_cabecalho = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col_idx in range(1, len(cabecalho) + 1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = fonte_cabecalho
        celula.fill = fundo_cabecalho

    for item in resultados:
        status = item.get("status", "")
        campos_a_exibir = {}
        for campo in colunas_campos:
            campos_a_exibir[campo] = "" if status == "erro" else item.get(campo, "")

        linha = [
            item.get("cliente", ""),
            item.get("ano", ""),
            item.get("arquivo", ""),
            item.get("caminho_completo", ""),
            item.get("origem", ""),
            status,
            item.get("erro_detalhe", ""),
        ]
        linha += [campos_a_exibir[campo] for campo in colunas_campos]
        ws.append(linha)

    for col_idx, titulo in enumerate(cabecalho, start=1):
        letra = get_column_letter(col_idx)
        largura = max(len(titulo) + 4, 14)
        if titulo in ("Caminho completo", "Erro Detalhe"):
            largura = 50
        ws.column_dimensions[letra].width = largura

    ws.freeze_panes = "A2"

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    log.info("Excel gerado em %s (%d linhas)", caminho_saida, len(resultados))
    return caminho_saida
